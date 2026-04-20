# Here is a **line‑by‑line explanation** of the Excel extraction code using `pandas` and `openpyxl`, presented as bullet points:

# - `import pandas as pd`
#   - Imports the `pandas` library and gives it the alias `pd`.
#   - Pandas is used for data manipulation, analysis, and reading structured data from Excel, CSV, etc.

# - `from openpyxl import load_workbook`
#   - Imports the `load_workbook` function from the `openpyxl` library.
#   - `openpyxl` is used for low‑level reading/writing of Excel `.xlsx` files, allowing access to cell formatting, formulas, and raw values.

# - `excel_path = 'C:\\Personal\\2024\\Learning\\Generative AI\\RAG\\27_Context_Engineering\\2_RAG\\1_Document_Processing\\Data\\sales_data.xlsx'`
#   - Assigns a string containing the file path to an Excel file named `sales_data.xlsx`.
#   - Double backslashes (`\\`) are used to escape backslashes in a regular Python string.

# - `# ── Method 1: pandas — best for data analysis ─────────────────────────────────`
#   - A comment indicating the start of the first method (using pandas for analysis).

# - `print("=" * 60)`
#   - Prints a line of 60 equals signs as a visual separator.

# - `print("METHOD 1: pandas read_excel (preserves table structure)")`
#   - Prints a descriptive header for the first method.

# - `print("=" * 60)`
#   - Prints another separator line.

# - `df = pd.read_excel(excel_path, sheet_name='Sales Data')`
#   - Uses `pd.read_excel()` to read the Excel file.
#   - The `sheet_name='Sales Data'` argument specifies which sheet to load (the sheet named “Sales Data”).
#   - The data is loaded into a pandas DataFrame called `df`.

# - `print(df.to_string(index=False))`
#   - Converts the entire DataFrame to a string representation.
#   - `to_string(index=False)` omits the row index numbers, printing only the data columns.
#   - The result is printed to the console.

# - `print(f"\nTotal Orders : {len(df)}")`
#   - Prints the number of rows in the DataFrame (total orders). `len(df)` gives the row count.

# - `print(f"Total Revenue: ${df['Total ($)'].sum():,.2f}")`
#   - Accesses the column named `'Total ($)'` from the DataFrame and computes the sum of its values.
#   - The sum is formatted as a currency with commas and two decimal places (`:,.2f`).
#   - The result is printed.

# - `print(f"\nRevenue by Category:")`
#   - Prints a header for the next output.

# - `print(df.groupby('Category')['Total ($)'].sum().to_string())`
#   - Groups the DataFrame by the `'Category'` column.
#   - For each group, selects the `'Total ($)'` column and computes the sum.
#   - The resulting Series is converted to a string (without index formatting) and printed.

# - `# ── Method 2: openpyxl — best for reading cell-by-cell with formatting ────────`
#   - A comment indicating the second method using openpyxl.

# - `print("\n" + "=" * 60)`
#   - Prints a newline followed by a separator line.

# - `print("METHOD 2: openpyxl (reads raw cell values + structure)")`
#   - Prints the header for the second method.

# - `print("=" * 60)`
#   - Prints another separator.

# - `wb = load_workbook(excel_path, data_only=True)  # data_only=True reads formula results`
#   - Loads the Excel workbook using `load_workbook`.
#   - `data_only=True` tells openpyxl to return the results of formulas instead of the formula strings. If `False`, it would return the formula itself.

# - `ws = wb['Sales Data']`
#   - Accesses the worksheet named `'Sales Data'` from the loaded workbook and assigns it to `ws`.

# - `# Print headers`
#   - A comment indicating that the following lines print the column headers.

# - `headers = [cell.value for cell in ws[1]]`
#   - Creates a list of header values by iterating over all cells in the first row (`ws[1]`). `cell.value` extracts the content of each cell.

# - `print(" | ".join(str(h) for h in headers))`
#   - Joins the header strings with a separator `" | "` (space‑pipe‑space) and prints the result.

# - `print("-" * 80)`
#   - Prints a line of 80 hyphens as a visual separator.

# - `# Print data rows (skip header + totals row)`
#   - A comment indicating that the loop prints the actual data rows, excluding the header row and a possible totals row at the end.

# - `for row in ws.iter_rows(min_row=2, max_row=ws.max_row - 1, values_only=True):`
#   - Iterates over rows in the worksheet.
#   - `min_row=2` starts from the second row (skip the header).
#   - `max_row=ws.max_row - 1` stops one row before the last row (assumes the last row contains totals).
#   - `values_only=True` returns the cell values instead of cell objects, improving performance.

# - `print(" | ".join(str(v) if v is not None else "" for v in row))`
#   - For each row, converts each cell value to a string (or an empty string if `None`) and joins them with `" | "`.
#   - The resulting string is printed.

# - `# ── Method 3: pandas — read all sheets at once ────────────────────────────────`
#   - A comment indicating the third method.

# - `print("\n" + "=" * 60)`
#   - Prints a newline followed by a separator line.

# - `print("METHOD 3: Read ALL sheets into a dictionary")`
#   - Prints the header for the third method.

# - `print("=" * 60)`
#   - Prints another separator.

# - `all_sheets = pd.read_excel(excel_path, sheet_name=None)`
#   - Reads the entire Excel file into a dictionary of DataFrames.
#   - `sheet_name=None` tells pandas to load all sheets; the dictionary keys are sheet names, and values are DataFrames.

# - `for sheet_name, sheet_df in all_sheets.items():`
#   - Loops over each sheet name and its corresponding DataFrame.

# - `print(f"\nSheet: '{sheet_name}' → {sheet_df.shape[0]} rows x {sheet_df.shape[1]} cols")`
#   - Prints the sheet name and its dimensions: number of rows (`shape[0]`) and number of columns (`shape[1]`).

# - `print(sheet_df.head(3).to_string(index=False))`
#   - Prints the first three rows of the sheet’s DataFrame.
#   - `head(3)` selects the first three rows, and `to_string(index=False)` prints them without row indices.

import pandas as pd
from openpyxl import load_workbook

excel_path = 'D:\GenAI Content\\AI code\\4_RAG_Indexing\\1_Document_Processing\\Data\\sales_data.xlsx'

# ── Method 1: pandas — best for data analysis ─────────────────────────────────
print("=" * 60)
print("METHOD 1: pandas read_excel (preserves table structure)")
print("=" * 60)

df = pd.read_excel(excel_path, sheet_name='Sales Data')
print(df.to_string(index=False))

# Basic analysis
print(f"\nTotal Orders : {len(df)}")
print(f"Total Revenue: ${df['Total ($)'].sum():,.2f}")
print(f"\nRevenue by Category:")
print(df.groupby('Category')['Total ($)'].sum().to_string())

# ── Method 2: openpyxl — best for reading cell-by-cell with formatting ────────
print("\n" + "=" * 60)
print("METHOD 2: openpyxl (reads raw cell values + structure)")
print("=" * 60)

wb = load_workbook(excel_path, data_only=True)  # data_only=True reads formula results
ws = wb['Sales Data']

# Print headers
headers = [cell.value for cell in ws[1]]
print(" | ".join(str(h) for h in headers))
print("-" * 80)

# Print data rows (skip header + totals row)
for row in ws.iter_rows(min_row=2, max_row=ws.max_row - 1, values_only=True):
    print(" | ".join(str(v) if v is not None else "" for v in row))

# ── Method 3: pandas — read all sheets at once ────────────────────────────────
print("\n" + "=" * 60)
print("METHOD 3: Read ALL sheets into a dictionary")
print("=" * 60)

all_sheets = pd.read_excel(excel_path, sheet_name=None)
for sheet_name, sheet_df in all_sheets.items():
    print(f"\nSheet: '{sheet_name}' → {sheet_df.shape[0]} rows x {sheet_df.shape[1]} cols")
    print(sheet_df.head(3).to_string(index=False))